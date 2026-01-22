# app/api/routes/calendar.py
from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import datetime, date
import io
import json
from typing import Any

from app.api.deps import get_db
from app.api.deps_auth import get_current_user, ensure_company_scope
from app.models.user import User

router = APIRouter(prefix="/companies", tags=["calendar"])


def _assert_can_modify_calendar(user: User):
    if user.role not in ("SYSADMIN", "COMPANY_ADMIN"):
        raise HTTPException(
            status_code=403,
            detail="Calendar can only be modified by SYSADMIN or COMPANY_ADMIN",
        )


def _norm_header(v: Any) -> str:
    # normalize "Fiscal   Day of theMonth" -> "fiscal day of themonth"
    s = str(v or "").strip().lower()
    s = " ".join(s.split())
    return s


def _as_int(v: Any, field: str, rownum: int) -> int:
    if v is None or str(v).strip() == "":
        raise HTTPException(status_code=400, detail=f"{field} missing at row {rownum}")
    try:
        # Excel sometimes gives floats for ints (e.g., 20260101.0)
        return int(float(str(v).strip()))
    except Exception:
        raise HTTPException(status_code=400, detail=f"{field} must be int at row {rownum}. Got: {v}")


def _as_date(v: Any, rownum: int) -> date:
    if v is None or str(v).strip() == "":
        raise HTTPException(status_code=400, detail=f"DateID missing at row {rownum}")

    # openpyxl date cells often come as datetime or date
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()

    # If someone gave YYYYMMDD as number/string
    if s.isdigit() and len(s) == 8:
        try:
            return datetime.strptime(s, "%Y%m%d").date()
        except Exception:
            raise HTTPException(status_code=400, detail=f"Bad DateID at row {rownum}: {v}")

    # Try ISO format YYYY-MM-DD
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        raise HTTPException(status_code=400, detail=f"Bad DateID at row {rownum}: {v}")


@router.post("/{company_id}/calendar/import")
async def import_calendar_excel(
    company_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_company_scope(user, company_id)
    _assert_can_modify_calendar(user)

    # 1) Ensure company exists
    c = db.execute(
        text("SELECT company_id, is_active, calendar_sheet_id FROM companies WHERE company_id = :cid"),
        {"cid": company_id},
    ).mappings().first()
    if not c:
        raise HTTPException(status_code=404, detail="Company not found")

    # 2) Read file bytes
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Empty file")

    # 3) Parse Excel
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl not installed on backend environment")

    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    ws = wb.active

    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="File must include header + at least 1 data row")

    # 4) Map headers -> column index (normalized)
    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        key = _norm_header(v)
        if key:
            headers[key] = col

    required = [
        "dateid",
        "fiscal year", "fiscal quarter", "fiscal month", "fiscal week", "fiscalyrwk",
        "fiscal day of the week", "fiscal day of themonth",
        "iso year", "iso quarter", "iso week", "iso month",
        "iso day of the week", "iso day of the month",
        "day name",
    ]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    # 5) Ensure / create the calendar sheet
    sheet = db.execute(
        text("""
            SELECT sheet_id
            FROM sheets
            WHERE company_id = :cid AND sheet_key = 'calendar'
            LIMIT 1
        """),
        {"cid": company_id},
    ).mappings().first()

    try:
        if sheet:
            sheet_id = int(sheet["sheet_id"])
            # Keep it active + refresh metadata
            db.execute(
                text("""
                    UPDATE sheets
                    SET sheet_name = 'Calendar',
                        description = COALESCE(description, 'Core calendar for fiscal/ISO'),
                        model_json = COALESCE(model_json, '{}'::jsonb) || CAST(:model_json AS jsonb),
                        is_active = true
                    WHERE company_id = :cid AND sheet_id = :sid
                """),
                {
                    "cid": company_id,
                    "sid": sheet_id,
                    "model_json": json.dumps({"type": "calendar", "source": "manual_upload"}),
                },
            )
        else:
            created = db.execute(
                text("""
                    INSERT INTO sheets (company_id, sheet_key, sheet_name, description, model_json, is_active)
                    VALUES (:cid, 'calendar', 'Calendar', 'Core calendar for fiscal/ISO', CAST(:model_json AS jsonb), true)
                    RETURNING sheet_id
                """),
                {
                    "cid": company_id,
                    "model_json": json.dumps({"type": "calendar", "source": "manual_upload"}),
                },
            ).mappings().one()
            sheet_id = int(created["sheet_id"])

        # 6) Replace existing calendar rows
        db.execute(text("DELETE FROM calendar WHERE company_id = :cid"), {"cid": company_id})

        # 7) Collect rows (bulk insert)
        payload_rows = []
        for r in range(2, ws.max_row + 1):
            dateid_raw = ws.cell(row=r, column=headers["dateid"]).value
            if dateid_raw is None or str(dateid_raw).strip() == "":
                continue

            date_id = _as_date(dateid_raw, r)

            payload_rows.append({
                "cid": company_id,
                "date_id": date_id,
                "fiscal_year": _as_int(ws.cell(row=r, column=headers["fiscal year"]).value, "Fiscal Year", r),
                "fiscal_quarter": _as_int(ws.cell(row=r, column=headers["fiscal quarter"]).value, "Fiscal Quarter", r),
                "fiscal_month": _as_int(ws.cell(row=r, column=headers["fiscal month"]).value, "Fiscal Month", r),
                "fiscal_week": _as_int(ws.cell(row=r, column=headers["fiscal week"]).value, "Fiscal Week", r),
                "fiscal_yrwk": _as_int(ws.cell(row=r, column=headers["fiscalyrwk"]).value, "FiscalYRWK", r),
                "fiscal_dow": _as_int(ws.cell(row=r, column=headers["fiscal day of the week"]).value, "Fiscal Day of the Week", r),
                "fiscal_dom": _as_int(ws.cell(row=r, column=headers["fiscal day of themonth"]).value, "Fiscal Day of theMonth", r),
                "iso_year": _as_int(ws.cell(row=r, column=headers["iso year"]).value, "ISO Year", r),
                "iso_quarter": _as_int(ws.cell(row=r, column=headers["iso quarter"]).value, "ISO Quarter", r),
                "iso_week": _as_int(ws.cell(row=r, column=headers["iso week"]).value, "ISO Week", r),
                "iso_month": _as_int(ws.cell(row=r, column=headers["iso month"]).value, "ISO Month", r),
                "iso_dow": _as_int(ws.cell(row=r, column=headers["iso day of the week"]).value, "ISO Day of the Week", r),
                "iso_dom": _as_int(ws.cell(row=r, column=headers["iso day of the month"]).value, "ISO Day of the Month", r),
                "day_name": (ws.cell(row=r, column=headers["day name"]).value or None),
            })

        if not payload_rows:
            raise HTTPException(status_code=400, detail="No calendar rows found in file")

        db.execute(
            text("""
                INSERT INTO calendar (
                  company_id, date_id,
                  fiscal_year, fiscal_quarter, fiscal_month, fiscal_week, fiscal_yrwk,
                  fiscal_dow, fiscal_dom,
                  iso_year, iso_quarter, iso_week, iso_month, iso_dow, iso_dom,
                  day_name
                )
                VALUES (
                  :cid, :date_id,
                  :fiscal_year, :fiscal_quarter, :fiscal_month, :fiscal_week, :fiscal_yrwk,
                  :fiscal_dow, :fiscal_dom,
                  :iso_year, :iso_quarter, :iso_week, :iso_month, :iso_dow, :iso_dom,
                  :day_name
                )
            """),
            payload_rows,
        )

        # 8) Tag company + activate
        db.execute(
            text("""
                UPDATE companies
                SET calendar_sheet_id = :sid,
                    is_active = true
                WHERE company_id = :cid
            """),
            {"cid": company_id, "sid": sheet_id},
        )

        db.commit()
        return {"ok": True, "calendar_sheet_id": sheet_id, "inserted": len(payload_rows)}

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Calendar import failed: {str(e)}")