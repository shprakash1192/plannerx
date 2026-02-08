from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import text
import openpyxl
import json
import io
import re

from app.db.session import get_db
from app.api.deps import require_company_admin_or_sysadmin

router = APIRouter(
    prefix="/companies/{company_id}/dimensions",
    tags=["dimensions"],
)

def _s(v) -> str:
    """Safe string conversion for Excel cells (handles int/float/etc)."""
    if v is None:
        return ""
    return str(v).strip()

def _to_bool(v, default=True):
    if v is None or str(v).strip() == "":
        return default
    s = str(v).strip().lower()
    if s in ("true", "t", "1", "yes", "y"):
        return True
    if s in ("false", "f", "0", "no", "n"):
        return False
    return default

def _norm_key(v) -> str:
    s = _s(v).lower()
    s = re.sub(r"\s+", "_", s)
    return s

def _norm_dtype(v) -> str:
    s = str(v or "").strip().upper()
    return s if s in ("TEXT", "NUMBER", "DATE") else "TEXT"

def _parse_json_obj(v):
    if v is None or str(v).strip() == "":
        return {}
    if isinstance(v, dict):
        return v
    s = str(v).strip()
    try:
        obj = json.loads(s)
    except Exception:
        raise ValueError("attributes_json must be valid JSON")
    if not isinstance(obj, dict):
        raise ValueError('attributes_json must be a JSON object like {"a":1}')
    return obj

def _norm_header(h: str) -> str:
    # "dimension_key*" -> "dimension_key"
    s = (h or "").strip().lower()
    s = s.replace("*", "")
    s = re.sub(r"\s+", "_", s)
    return s

@router.post("/import")
def import_dimensions_excel(
    company_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _user=Depends(require_company_admin_or_sysadmin),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx/.xlsm file")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {str(e)}")

    # Accept either "Values" or "DimensionValues"
    dim_sheet = "Dimensions"
    val_sheet = "Values" if "Values" in wb.sheetnames else ("DimensionValues" if "DimensionValues" in wb.sheetnames else None)

    if dim_sheet not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Workbook must contain a sheet named 'Dimensions'")
    if not val_sheet:
        raise HTTPException(status_code=400, detail="Workbook must contain a sheet named 'Values' (or 'DimensionValues')")

    dim_ws = wb[dim_sheet]
    val_ws = wb[val_sheet]

    def read_sheet(ws):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], {}

        header_raw = [str(c or "").strip() for c in rows[0]]
        header = [_norm_header(h) for h in header_raw]
        idx = {h: i for i, h in enumerate(header) if h}

        data = []
        for r in rows[1:]:
            if all((c is None or str(c).strip() == "") for c in r):
                continue
            data.append(r)

        return data, idx

    dim_rows, dim_idx = read_sheet(dim_ws)
    val_rows, val_idx = read_sheet(val_ws)

    required_dim_cols = ["dimension_key", "dimension_name"]
    required_val_cols = ["dimension_key", "value_key", "value_name"]

    for c in required_dim_cols:
        if c not in dim_idx:
            raise HTTPException(status_code=400, detail=f"'Dimensions' sheet missing column: {c}")
    for c in required_val_cols:
        if c not in val_idx:
            raise HTTPException(status_code=400, detail=f"'{val_sheet}' sheet missing column: {c}")

    summary = {
        "dimensions": {"created": 0, "updated": 0, "skipped": 0, "errors": []},
        "values": {"created": 0, "updated": 0, "skipped": 0, "errors": []},
        "values_sheet": val_sheet,
    }

    try:
        # ---- DIMENSIONS UPSERT ----
        for n, r in enumerate(dim_rows, start=2):
            try:
                dkey = _norm_key(r[dim_idx["dimension_key"]])
                dname = _s(r[dim_idx["dimension_name"]])
                if not dkey or not dname:
                    summary["dimensions"]["skipped"] += 1
                    continue

                dtype = _norm_dtype(r[dim_idx["data_type"]]) if "data_type" in dim_idx else "TEXT"
                desc = r[dim_idx["description"]] if "description" in dim_idx else None
                desc = None if _s(desc) == "" else _s(desc)
                is_active = _to_bool(r[dim_idx["is_active"]], True) if "is_active" in dim_idx else True

                existing = db.execute(
                    text("""
                        SELECT dimension_id
                        FROM dimensions
                        WHERE company_id = :cid AND dimension_key = :dkey
                    """),
                    {"cid": company_id, "dkey": dkey},
                ).fetchone()

                if existing:
                    db.execute(
                        text("""
                            UPDATE dimensions
                            SET
                              dimension_name = :name,
                              description    = :desc,
                              is_active      = :is_active
                            WHERE company_id = :cid AND dimension_key = :dkey
                        """),
                        {"cid": company_id, "dkey": dkey, "name": dname, "desc": desc, "is_active": is_active},
                    )
                    summary["dimensions"]["updated"] += 1
                else:
                    db.execute(
                        text("""
                            INSERT INTO dimensions (company_id, dimension_key, dimension_name, description, data_type, is_active)
                            VALUES (:cid, :dkey, :name, :desc, :dtype, :is_active)
                        """),
                        {"cid": company_id, "dkey": dkey, "name": dname, "desc": desc, "dtype": dtype, "is_active": is_active},
                    )
                    summary["dimensions"]["created"] += 1

            except Exception as e:
                summary["dimensions"]["errors"].append({"row": n, "error": str(e)})

        # ---- VALUES UPSERT ----
        dim_map = dict(
            db.execute(
                text("""
                    SELECT dimension_key, dimension_id
                    FROM dimensions
                    WHERE company_id = :cid
                """),
                {"cid": company_id},
            ).fetchall()
        )

        for n, r in enumerate(val_rows, start=2):
            try:
                dkey = _norm_key(r[val_idx["dimension_key"]])
                vkey = _norm_key(r[val_idx["value_key"]])
                vname = _s(r[val_idx["value_name"]])

                if not dkey or not vkey or not vname:
                    summary["values"]["skipped"] += 1
                    continue

                if dkey not in dim_map:
                    raise ValueError(
                        f"dimension_key '{dkey}' not found for this company. "
                        f"Make sure it exists in Dimensions sheet and matches exactly after normalization."
                    )

                did = dim_map[dkey]

                sort_order = None
                if "sort_order" in val_idx:
                    raw = r[val_idx["sort_order"]]
                    if _s(raw) != "":
                        sort_order = int(float(raw))

                is_active = _to_bool(r[val_idx["is_active"]], True) if "is_active" in val_idx else True
                attrs = _parse_json_obj(r[val_idx["attributes_json"]]) if "attributes_json" in val_idx else {}

                existing = db.execute(
                    text("""
                        SELECT dimension_value_id
                        FROM dimension_values
                        WHERE company_id = :cid AND dimension_id = :did AND value_key = :vkey
                    """),
                    {"cid": company_id, "did": did, "vkey": vkey},
                ).fetchone()

                if existing:
                    db.execute(
                        text("""
                            UPDATE dimension_values
                            SET
                              value_name      = :name,
                              sort_order      = :sort_order,
                              attributes_json = :attrs,
                              is_active       = :is_active
                            WHERE company_id = :cid AND dimension_id = :did AND value_key = :vkey
                        """),
                        {"cid": company_id, "did": did, "vkey": vkey, "name": vname, "sort_order": sort_order, "attrs": json.dumps(attrs), "is_active": is_active},
                    )
                    summary["values"]["updated"] += 1
                else:
                    db.execute(
                        text("""
                            INSERT INTO dimension_values (company_id, dimension_id, value_key, value_name, sort_order, attributes_json, is_active)
                            VALUES (:cid, :did, :vkey, :name, :sort_order, :attrs, :is_active)
                        """),
                        {"cid": company_id, "did": did, "vkey": vkey, "name": vname, "sort_order": sort_order, "attrs": json.dumps(attrs), "is_active": is_active},
                    )
                    summary["values"]["created"] += 1

            except Exception as e:
                summary["values"]["errors"].append({"row": n, "error": str(e)})

        print("IMPORT SUMMARY:", json.dumps(summary, indent=2))
        db.commit()
        return {"ok": True, "summary": summary}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")